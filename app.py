import os
import pyaudio
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, jsonify
from threading import Thread
from transformers import pipeline
import speech_recognition as sr
from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"

# Configuration
OUTPUT_FILE = "data/transcriptions.xlsx"
LANGUAGE_CODE = "en-US"
app = Flask(__name__)
transcriptions, sentiments, tones, feedbacks = [], [], [], []

# Load tokenizer and model
tokenizer = AutoTokenizer.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")
model = AutoModelForSequenceClassification.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")

# Create a pipeline for sentiment analysis
sentiment_analyzer = pipeline("sentiment-analysis", model=model, tokenizer=tokenizer)
tone_analyzer = pipeline("text-classification", model="j-hartmann/emotion-english-distilroberta-base")

# Feedback logic
def provide_feedback(sentiment, tone):
    feedback = f"Feedback: Sentiment is {sentiment} with tone {tone}. "
    if "negative" in sentiment.lower():
        if "angry" in tone.lower():
            feedback += "Consider calming the situation or rephrasing. Offer immediate resolution."
        elif "sad" in tone.lower():
            feedback += "Empathize with the customer and provide a reassuring response."
        elif "fearful" in tone.lower():
            feedback += "Address the concerns and reassure the customer with a detailed explanation."
        else:
            feedback += "Address the customer's concerns promptly and offer assistance."
    
    elif "positive" in sentiment.lower():
        if "happy" in tone.lower():
            feedback += "Buyer is engaged. Keep up the positive flow."
        elif "excited" in tone.lower():
            feedback += "Customer is thrilled. Consider suggesting additional products or upgrades."
        elif "relaxed" in tone.lower():
            feedback += "The customer is satisfied. Maintain a supportive tone."
        else:
            feedback += "Continue delivering excellent service to reinforce positive engagement."
    
    elif "neutral" in sentiment.lower():
        if "bored" in tone.lower():
            feedback += "Reignite the customer's interest with engaging details or promotions."
        elif "uncertain" in tone.lower():
            feedback += "Clarify any doubts and provide additional information."
        else:
            feedback += "Maintain the current approach, ensuring clarity and support."

    else:
        feedback += "No specific advice for this combination. Continue monitoring the interaction."

    print(feedback)
    return feedback


# Audio transcription
def transcribe_audio():
    global transcriptions, sentiments, tones, feedbacks
    recognizer = sr.Recognizer()
def transcribe_audio():
    global transcriptions, sentiments, tones, feedbacks
    recognizer = sr.Recognizer()

    while True:
        with sr.Microphone() as source:
            print("Listening...")
            try:
                audio_data = recognizer.listen(source, timeout=5)
                print("Recognizing...")
                text = recognizer.recognize_google(audio_data, language=LANGUAGE_CODE)
                print(f"Transcribed: {text}")
                transcriptions.append(text)

                # Analyze sentiment and tone
                sentiment = analyze_sentiment(text)
                tone = analyze_tone(text)
                sentiments.append(sentiment)
                tones.append(tone)

                # Provide feedback and store it
                feedback = provide_feedback(sentiment, tone)
                feedbacks.append(feedback)

                # Save to Excel
                save_to_excel(text, sentiment, tone, feedback)

            except sr.UnknownValueError:
                print("Could not understand the audio.")
                feedback = "Error: Could not understand the audio."
                feedbacks.append(feedback)
            except sr.RequestError as e:
                print(f"Request error: {e}")
                feedback = f"Error: {e}"
                feedbacks.append(feedback)
            except Exception as e:
                print(f"An error occurred: {e}")
                feedback = f"Error: {e}"
                feedbacks.append(feedback)


# Sentiment analysis
def analyze_sentiment(text):
    try:
        result = sentiment_analyzer(text)[0]
        return result["label"]
    except Exception as e:
        print(f"Error in sentiment analysis: {e}")
        return "UNKNOWN"

# Tone analysis
def analyze_tone(text):
    try:
        result = tone_analyzer(text)[0]
        return result["label"]
    except Exception as e:
        print(f"Error in tone analysis: {e}")
        return "UNKNOWN"

# Save to Excel
def save_to_excel(text, sentiment, tone, feedback):
    if not os.path.exists(OUTPUT_FILE):
        os.makedirs("data", exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Transcription", "Sentiment", "Tone", "Feedback"])
        workbook.save(OUTPUT_FILE)

    workbook = load_workbook(OUTPUT_FILE)
    sheet = workbook.active
    sheet.append([text, sentiment, tone, feedback])
    workbook.save(OUTPUT_FILE)

# Flask routes
@app.route("/")
def dashboard():
    return render_template("dashboard.html", data=zip(transcriptions, sentiments, tones, feedbacks))

@app.route("/api/data")
def api_data():
    return jsonify({"transcriptions": transcriptions, "sentiments": sentiments, "tones": tones,  "feedbacks": feedbacks})

# Start transcription in a thread
def start_transcription():
    transcribe_audio()

# Run the Flask app
if __name__ == "__main__":
    transcription_thread = Thread(target=start_transcription)
    transcription_thread.start()
    app.run(host="0.0.0.0", port=5000)